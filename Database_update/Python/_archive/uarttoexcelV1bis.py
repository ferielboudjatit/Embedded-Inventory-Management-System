import serial
import re
import schedule
import time
import logging
from openpyxl import load_workbook
from openpyxl.packaging.custom import StringProperty, CustomPropertyList
import xlsxwriter
import shutil
import os
import time


path = "C:\\Users\\boudjatf\\Downloads\\Database\\Database-V2.xlsx"

def is_file_available(filepath):
    try:
        with open(filepath, 'a'):
            return True
    except (IOError, OSError, PermissionError):
        return False

def configure_excelfile():
    workbook = load_workbook(path)
    company_guid = "23add6c0-cfdb-4bb9-b90f-bf23b83aa6c0"
    name = company_guid
    site_id = "75e027c9-20d5-47d5-b82f-77d7cd041e8f"
    action_id = "15e5cbb6-2a0b-4ef6-9619-008dbbadc8e9"

    workbook.custom_doc_props = CustomPropertyList()
    
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_Enabled", value="true"))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_SetDate", value="2022-04-07T13:27:25Z"))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_Method", value="Privileged"))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_Name", value=name))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_SiteId", value=site_id))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_ActionId", value=action_id))
    workbook.custom_doc_props.append(StringProperty(name=f"MSIP_Label_{company_guid}_ContentBits", value="2"))


    temp_path = path + ".tmp"

    workbook.save(temp_path)
    workbook.close()
    
    
    if is_file_available(path):
        shutil.move(temp_path, path)
        print(f"File successfully updated: {path}")
    else:
        
        base_name = os.path.splitext(path)[0]
        extension = os.path.splitext(path)[1]
        
        version = 1
        while True:
            new_path = f"{base_name}_v{version:03d}{extension}"
            if is_file_available(new_path):
                try:
                    shutil.move(temp_path, new_path)
                    print(f"File saved as: {new_path}")
                    break
                except (IOError, OSError, PermissionError):
                    
                    version += 1
                    continue
            else:
                version += 1
                
            if version > 999:
                print("Error: Unable to find available filename after 999 attempts")
                os.remove(temp_path)  
                break


def update_spreadsheet(user_uid, board_uid, stock_state):
    #print(wb.sheetnames)  
    wb = load_workbook(path)
    sheet1 = wb['NUCLEO']
    sheet2 = wb['Users']
    sheet3 = wb['Roaming']
    


    board_name = None
    for row in sheet1.iter_rows(min_row=2):  
        if row[4].value == board_uid: 
            row[5].value = 'Y' if stock_state == 1 else 'N'  
            board_name = row[2].value
            break
    print("Board name",board_name)

    user_name = None
    for row in sheet2.iter_rows(min_row=2):  
        if str(row[1].value) == str(user_uid):
            user_name = row[0].value  
            break
    if user_name != None:
        print("Username",user_name) 

    
    for row in sheet1.iter_rows(min_row=2):  
        if row[4].value == board_uid:  
            row[6].value = user_name if stock_state == 2 else None
            break

    #if user_name and board_name:
    for row in sheet3.iter_rows(min_row=2):
#            for column in sheet3.iter_cols(min_col=2):
        #print("row values", row[0].value)
        if str(row[1].value) == str(board_name):
            for idx, cell in enumerate(row[2:], start=2): 
                column_header = sheet3.cell(row=1, column=idx+1).value
                #print("column header", column_header)
                if stock_state == 2:
                    #print("entered stock_state == 2")
                    if column_header == user_name:
                        #print("entered column_header == user_name")
                        cell.value = 1
                elif stock_state == 1:
                        #print("cell", cell.value)
                        if cell.value == 1:
                            cell.value = 0 
                #break
    configure_excelfile()



#logging.basicConfig(level=logging.INFO, format='%(message)s', handlers=[logging.StreamHandler()])
ser = serial.Serial('COM9', 115200, timeout=1)
pattern_in = r"Received Buffer: Message board got into stock : ([0-9A-F]+)_(\d)"
pattern_out = r"Received Buffer: Message board got out of stock : ([0-9A-F]+)_([0-9A-F]+)_(\d)"

try:
    while ser.is_open:
        #user_uid = None
        #board_uid = None
        #stock_state = None
        message = ser.readline().decode('utf-8').strip()
        if message :
            match_in = re.match(pattern_in, message)
            match_out = re.match(pattern_out, message)
            print(message)
            if match_in:
                user_uid = None
                board_uid = match_in.group(1)
                stock_state = int(match_in.group(2))
                print(f"Board UID: {board_uid}, Stock State: {stock_state}")
                update_spreadsheet(user_uid = None, board_uid=board_uid, stock_state=stock_state)
            elif match_out:
                user_uid = match_out.group(1)
                board_uid = match_out.group(2)
                stock_state = int(match_out.group(3))
                print(f"User UID: {user_uid}, Board UID: {board_uid}, Stock State: {stock_state}")
                update_spreadsheet(user_uid = user_uid, board_uid = board_uid, stock_state = stock_state)

except serial.SerialException as e:
    print(f"Error reading from serial port: {e}")
finally:
    if ser.is_open:
        ser.close()
        print("Serial port closed.")


# while (1):
#     ser.open()
#     message = ser.readline().decode('utf-8').strip()
#     if message :
#         match_in = re.match(pattern_in, message)
#         match_out = re.match(pattern_out, message)
#         print(message)
#         if match_in:
#             user_uid = None
#             board_uid = match_in.group(1)
#             stock_state = int(match_in.group(2))
#             print(f"Board UID: {board_uid}, Stock State: {stock_state}")
#             # update_spreadsheet(user_uid = None, board_uid=board_uid, stock_state=stock_state)
#         elif match_out:
#             user_uid = match_out.group(1)
#             board_uid = match_out.group(2)
#             stock_state = int(match_out.group(3))
#             print(f"User UID: {user_uid}, Board UID: {board_uid}, Stock State: {stock_state}")
    
#         update_spreadsheet(user_uid = user_uid, board_uid = board_uid, stock_state = stock_state)
#     ser.close()