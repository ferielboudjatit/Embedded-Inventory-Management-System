import serial
import re
import schedule
import time
import logging
from openpyxl import load_workbook
import xlsxwriter

path = "C:\\Users\\boudjatf\\Downloads\\Database\\Database_V3.xlsx"

def configure_excelfile():
    workbook = xlsxwriter.Workbook(path)
    company_guid = "23add6c0-cfdb-4bb9-b90f-bf23b83aa6c0"
    name = company_guid
    site_id = "75e027c9-20d5-47d5-b82f-77d7cd041e8f"
    action_id = "15e5cbb6-2a0b-4ef6-9619-008dbbadc8e9"

    workbook.set_custom_property(f"MSIP_Label_{company_guid}_Enabled", "true", "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_SetDate", "2022-04-07T13:27:25Z", "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_Method", "Privileged", "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_Name", name, "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_SiteId", site_id, "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_ActionId", action_id, "text")
    workbook.set_custom_property(f"MSIP_Label_{company_guid}_ContentBits", "2", "text")

    workbook.close()

def update_spreadsheet(user_uid, board_uid, stock_state):
    #print(wb.sheetnames)  
    wb = load_workbook(path)
    sheet1 = wb['NUCLEO']
    sheet2 = wb['Users']
    sheet3 = wb['Roaming']
    


    board_name = None
    for row in sheet1.iter_rows(min_row=2):  
        if row[4].value == board_uid: 
            row[5].value = 'Y' if stock_state == 1 else 'N'  
            board_name = row[2].value
            break
    print("Board name",board_name)

    user_name = None
    for row in sheet2.iter_rows(min_row=2):  
        if str(row[1].value) == str(user_uid):
            user_name = row[0].value  
            break
    if user_name != None:
        print("Username",user_name) 

    
    for row in sheet1.iter_rows(min_row=2):  
        if row[4].value == board_uid:  
            row[6].value = user_name if stock_state == 2 else None
            break

    #if user_name and board_name:
    for row in sheet3.iter_rows(min_row=2):
#            for column in sheet3.iter_cols(min_col=2):
        #print("row values", row[0].value)
        if str(row[1].value) == str(board_name):
            for idx, cell in enumerate(row[2:], start=2): 
                column_header = sheet3.cell(row=1, column=idx+1).value
                #print("column header", column_header)
                if stock_state == 2:
                    #print("entered stock_state == 2")
                    if column_header == user_name:
                        #print("entered column_header == user_name")
                        cell.value = 1
                elif stock_state == 1:
                        #print("cell", cell.value)
                        if cell.value == 1:
                            cell.value = 0 
                #break
    configure_excelfile()
    wb.save(path)

#logging.basicConfig(level=logging.INFO, format='%(message)s', handlers=[logging.StreamHandler()])
ser = serial.Serial('COM9', 115200, timeout=1)
pattern_in = r"Received Buffer: Message board got into stock : ([0-9A-F]+)_(\d)"
pattern_out = r"Received Buffer: Message board got out of stock : ([0-9A-F]+)_([0-9A-F]+)_(\d)"
try:
    while ser.is_open:
        #user_uid = None
        #board_uid = None
        #stock_state = None
        message = ser.readline().decode('utf-8').strip()
        if message :
            match_in = re.match(pattern_in, message)
            match_out = re.match(pattern_out, message)
            print(message)
            if match_in:
                user_uid = None
                board_uid = match_in.group(1)
                stock_state = int(match_in.group(2))
                print(f"Board UID: {board_uid}, Stock State: {stock_state}")
                update_spreadsheet(user_uid = None, board_uid=board_uid, stock_state=stock_state)
            elif match_out:
                user_uid = match_out.group(1)
                board_uid = match_out.group(2)
                stock_state = int(match_out.group(3))
                print(f"User UID: {user_uid}, Board UID: {board_uid}, Stock State: {stock_state}")
            #else:
                #break
        #else:
            #time.sleep(10000)
            #break
                update_spreadsheet(user_uid = user_uid, board_uid = board_uid, stock_state = stock_state)

except serial.SerialException as e:
    print(f"Error reading from serial port: {e}")
finally:
    if ser.is_open:
        ser.close()
        print("Serial port closed.")

'''
def job():
    pattern = r"Message board got (into|out of) stock : ([0-9A-F]+)_([0-9A-F]+)_(\d)"
    ser = serial.Serial('COM15', 115200, timeout=1)
    try:
        while ser.is_open:
            message = ser.readline().decode('utf-8').strip()
            print(message)
            match = re.match(pattern, message)
            if match:
                user_uid = match.group(2)
                board_uid = match.group(3)
                stock_state = int(match.group(4))
                print(f"User UID: {user_uid}, Board UID: {board_uid}, Stock State: {stock_state}")
                update_spreadsheet(user_uid, board_uid, stock_state)
    except serial.SerialException as e:
        print(f"Error reading from serial port: {e}")
    finally:
        if ser.is_open:
            ser.close()
            print("Serial port closed.")

schedule.every(10).seconds.do(job)  

while True:
    schedule.run_pending()
    time.sleep(1)
'''